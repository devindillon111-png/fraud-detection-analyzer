import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os
from datetime import datetime

# --- File Paths ---
script_dir = os.path.dirname(os.path.abspath(__file__))
flagged_path = os.path.join(script_dir, "..", "outputs", "flagged_transactions.csv")
full_path = os.path.join(script_dir, "..", "data", "transactions.csv")
output_path = os.path.join(script_dir, "..", "outputs", "fraud_detection_report.xlsx")

# --- Load Data ---
flagged_df = pd.read_csv(flagged_path)
full_df = pd.read_csv(full_path)

# Convert boolean columns to numeric
bool_cols = ['Is_Weekend', 'Is_Round_Number', 'Is_Threshold', 'Missing_Vendor', 'Missing_Employee', 'ML_Anomaly']
for col in bool_cols:
    if col in flagged_df.columns:
        flagged_df[col] = flagged_df[col].map({True: 1, False: 0, 'True': 1, 'False': 0}).fillna(0).astype(int)


# --- Create Workbook ---
wb = openpyxl.Workbook()

# --- Color Palette ---
RED = "C00000"
DARK_GRAY = "2F2F2F"
LIGHT_GRAY = "F2F2F2"
ORANGE = "FF6600"
WHITE = "FFFFFF"
YELLOW = "FFD966"

# ============================================================
# SHEET 1: EXECUTIVE SUMMARY
# ============================================================
ws1 = wb.active
ws1.title = "Executive Summary"

# Header
ws1.merge_cells("A1:F1")
ws1["A1"] = "FRAUD DETECTION ANALYZER — EXECUTIVE SUMMARY"
ws1["A1"].font = Font(bold=True, size=16, color=WHITE)
ws1["A1"].fill = PatternFill("solid", fgColor=DARK_GRAY)
ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:F2")
ws1["A2"] = f"Report Generated: {datetime.now().strftime('%B %d, %Y')}"
ws1["A2"].font = Font(italic=True, size=11, color=DARK_GRAY)
ws1["A2"].alignment = Alignment(horizontal="center")

# KPI Section Title
ws1.merge_cells("A4:F4")
ws1["A4"] = "KEY METRICS"
ws1["A4"].font = Font(bold=True, size=12, color=WHITE)
ws1["A4"].fill = PatternFill("solid", fgColor=RED)
ws1["A4"].alignment = Alignment(horizontal="center")

# KPI Data
total = len(full_df)
flagged = len(flagged_df)
flag_rate = round(flagged / total * 100, 1)
total_risk = round(flagged_df['Amount'].sum(), 2)
avg_flagged = round(flagged_df['Amount'].mean(), 2)
ml_anomalies = int(flagged_df['ML_Anomaly'].sum()) if 'ML_Anomaly' in flagged_df.columns else 0

kpis = [
    ("Total Transactions Analyzed", f"{total:,}"),
    ("Total Flagged Transactions", f"{flagged:,}"),
    ("Overall Flag Rate", f"{flag_rate}%"),
    ("Total Dollar Amount at Risk", f"${total_risk:,.2f}"),
    ("Average Flagged Transaction Amount", f"${avg_flagged:,.2f}"),
    ("ML Anomalies Detected", f"{ml_anomalies:,}"),
]

for i, (label, value) in enumerate(kpis):
    row = i + 5
    ws1[f"A{row}"] = label
    ws1[f"A{row}"].font = Font(bold=True, size=11)
    ws1[f"A{row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws1[f"B{row}"] = value
    ws1[f"B{row}"].font = Font(bold=True, size=11, color=RED)
    ws1[f"B{row}"].alignment = Alignment(horizontal="center")

# Department Summary
ws1.merge_cells("A12:F12")
ws1["A12"] = "FLAGS BY DEPARTMENT"
ws1["A12"].font = Font(bold=True, size=12, color=WHITE)
ws1["A12"].fill = PatternFill("solid", fgColor=RED)
ws1["A12"].alignment = Alignment(horizontal="center")

dept_summary = flagged_df.groupby('Department').agg(
    Flagged_Count=('Amount', 'count'),
    Total_Amount=('Amount', 'sum')
).reset_index().sort_values('Flagged_Count', ascending=False)

headers = ["Department", "Flagged Transactions", "Total Amount at Risk"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=13, column=col, value=header)
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=DARK_GRAY)
    cell.alignment = Alignment(horizontal="center")

for i, row in dept_summary.iterrows():
    r = 14 + list(dept_summary.index).index(i)
    ws1.cell(row=r, column=1, value=row['Department'])
    ws1.cell(row=r, column=2, value=int(row['Flagged_Count'])).alignment = Alignment(horizontal="center")
    ws1.cell(row=r, column=3, value=f"${row['Total_Amount']:,.2f}").alignment = Alignment(horizontal="center")
    if r % 2 == 0:
        for col in range(1, 4):
            ws1.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

ws1.column_dimensions['A'].width = 35
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 25

# ============================================================
# SHEET 2: FLAGGED TRANSACTIONS
# ============================================================
ws2 = wb.create_sheet("Flagged Transactions")

ws2.merge_cells("A1:N1")
ws2["A1"] = "FLAGGED TRANSACTIONS — DETAILED VIEW"
ws2["A1"].font = Font(bold=True, size=14, color=WHITE)
ws2["A1"].fill = PatternFill("solid", fgColor=DARK_GRAY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

display_cols = ['Date', 'Transaction_ID', 'Amount', 'Department', 'Vendor_Name',
                'Employee_ID', 'Is_Weekend', 'Is_Round_Number', 'Is_Threshold',
              'Missing_Vendor', 'Rule_Flag_Count', 'ML_Anomaly', 'Risk_Score', 'Risk_Level']
available_cols = [c for c in display_cols if c in flagged_df.columns]

for col, header in enumerate(available_cols, 1):
    cell = ws2.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=RED)
    cell.alignment = Alignment(horizontal="center")

for row_idx, row in flagged_df[available_cols].iterrows():
    excel_row = row_idx + 3
    for col_idx, value in enumerate(row, 1):
        cell = ws2.cell(row=excel_row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center")
        if excel_row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        # Highlight high flag count rows
        if col_idx == len(available_cols) and isinstance(value, (int, float)) and value >= 3:
            for c in range(1, len(available_cols) + 1):
                ws2.cell(row=excel_row, column=c).fill = PatternFill("solid", fgColor=YELLOW)

for col in range(1, len(available_cols) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 20

# ============================================================
# SHEET 3: FLAG SUMMARY
# ============================================================
ws3 = wb.create_sheet("Flag Summary")

ws3.merge_cells("A1:C1")
ws3["A1"] = "FLAG TYPE SUMMARY"
ws3["A1"].font = Font(bold=True, size=14, color=WHITE)
ws3["A1"].fill = PatternFill("solid", fgColor=DARK_GRAY)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

flag_cols = {
    'Is_Weekend': 'Weekend Transaction',
    'Is_Round_Number': 'Round Number',
    'Is_Threshold': 'Threshold Manipulation',
    'Missing_Vendor': 'Missing Vendor',
    'ML_Anomaly': 'ML Anomaly'
}

headers = ["Flag Type", "Count", "% of Flagged Transactions"]
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=2, column=col, value=header)
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=RED)
    cell.alignment = Alignment(horizontal="center")

for i, (col_name, label) in enumerate(flag_cols.items()):
    row = i + 3
    count = int(pd.to_numeric(flagged_df[col_name], errors='coerce').fillna(0).sum()) if col_name in flagged_df.columns else 0
    pct = round(count / flagged * 100, 1)
    ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
    ws3.cell(row=row, column=2, value=count).alignment = Alignment(horizontal="center")
    ws3.cell(row=row, column=3, value=f"{pct}%").alignment = Alignment(horizontal="center")
    if row % 2 == 0:
        for c in range(1, 4):
            ws3.cell(row=row, column=c).fill = PatternFill("solid", fgColor=LIGHT_GRAY)

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 30

# --- Save ---
wb.save(output_path)
print(f"Report saved to: {output_path}")
print("Excel report generation complete!")